import os
import base64
import re
from openpyxl import Workbook, load_workbook
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from google.auth.transport.requests import Request
from dotenv import load_dotenv

load_dotenv()

SCOPES = ['https://www.googleapis.com/auth/gmail.modify']
ANEXOS_DIR = "anexos"
PLANILHA = "candidatos.xlsx"
os.makedirs(ANEXOS_DIR, exist_ok=True)

CREDENTIALS_FILE = os.getenv("GOOGLE_CREDENTIALS", "credentials.json")
TOKEN_FILE = os.getenv("GOOGLE_TOKEN", "token.json")

# Função para destacar alertas visuais no terminal (vermelho)
def alerta_visual(msg):
    print(f"\033[91m{msg}\033[0m")  # ANSI vermelho

def autenticar_gmail():
    creds = None
    if os.path.exists(TOKEN_FILE):
        creds = Credentials.from_authorized_user_file(TOKEN_FILE, SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(CREDENTIALS_FILE, SCOPES)
            creds = flow.run_local_server(port=0)
        with open(TOKEN_FILE, "w") as token:
            token.write(creds.to_json())
    return creds

def extrair_dados(corpo, assunto):
    nome_match = re.search(r"(Nome completo|Nome):\s*(.*?)\n", corpo, re.IGNORECASE)
    telefone_match = re.search(r"Telefone:\s*(.*?)\n", corpo)
    vaga_assunto_match = re.search(r"(?:Candidatura\s*-\s*|vaga de|para a vaga|cargo de)\s*(.*)", assunto, re.IGNORECASE)

    nome = nome_match.group(2).strip() if nome_match and len(nome_match.groups()) >= 2 else "Não encontrado"
    telefone = telefone_match.group(1).strip() if telefone_match and len(telefone_match.groups()) >= 1 else "Não encontrado"
    vaga = vaga_assunto_match.group(1).strip() if vaga_assunto_match else "Não encontrada"

    if vaga == "Não encontrada":
        vaga_corpo_match = re.search(r"(?:vaga de|para a vaga|cargo de)\s*([^\.\n]*)", corpo, re.IGNORECASE)
        vaga = vaga_corpo_match.group(1).strip() if vaga_corpo_match else "Não encontrada"

    return {
        "nome": nome,
        "telefone": telefone,
        "vaga": vaga
    }

def registrar_planilha(dados, anexo):
    if os.path.exists(PLANILHA):
        wb = load_workbook(PLANILHA)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Nome", "Telefone", "Vaga", "Anexo"])
    ws.append([dados["nome"], dados["telefone"], dados["vaga"], anexo or "SEM ANEXO"])
    wb.save(PLANILHA)

def salvar_anexo(service, msg_id, part):
    if 'body' in part and 'attachmentId' in part['body']:
        att_id = part['body']['attachmentId']
        att = service.users().messages().attachments().get(userId="me", messageId=msg_id, id=att_id).execute()
        data = base64.urlsafe_b64decode(att['data'].encode('UTF-8'))
        filename = part['filename'] or "anexo.pdf"
        filepath = os.path.join(ANEXOS_DIR, filename)
        with open(filepath, "wb") as f:
            f.write(data)
        return filepath
    return None

def obter_corpo(part):
    if part.get("mimeType") == "text/plain" and 'data' in part['body']:
        return base64.urlsafe_b64decode(part["body"]["data"]).decode("utf-8")
    elif part.get("mimeType") == "text/html" and 'data' in part['body']:
        return base64.urlsafe_b64decode(part["body"]["data"]).decode("utf-8")
    elif part.get("parts"):
        for sub_part in part["parts"]:
            corpo = obter_corpo(sub_part)
            if corpo:
                return corpo
    return None

def processar_emails(service):
    # Busca apenas e-mails não lidos (is:unread) e limita a 50 resultados por performance
    results = service.users().messages().list(userId="me", maxResults=50, q="is:unread").execute()
    mensagens = results.get("messages", [])

    print(f"E-mails não lidos encontrados: {len(mensagens)}")

    for msg in mensagens:
        msg_data = service.users().messages().get(userId="me", id=msg["id"]).execute()
        payload = msg_data["payload"]
        headers = payload["headers"]
        assunto = next((h["value"] for h in headers if h["name"] == "Subject"), "(Sem Assunto)")

        corpo = obter_corpo(payload) or ""
        dados = extrair_dados(corpo, assunto)

        # Filtra apenas os com dados relevantes
        if dados["nome"] == "Não encontrado" or dados["telefone"] == "Não encontrado" or not dados["vaga"] or dados["vaga"] == "Não encontrada":
            continue

        print(f"\n Processando: {assunto}")
        print(f"-> Dados extraídos: {dados}")

        anexo_salvo = None
        for part in payload.get("parts", []):
            if "filename" in part and part["filename"]:
                anexo_salvo = salvar_anexo(service, msg["id"], part)
                print(f"-> Anexo salvo em: {anexo_salvo}")

        if not anexo_salvo:
            alerta_visual("⚠️  Nenhum anexo encontrado neste e-mail!")

        registrar_planilha(dados, anexo_salvo)
        print("Registro adicionado na planilha.")

        # Marca como lido após o processamento
        service.users().messages().modify(userId="me", id=msg["id"], body={
            "removeLabelIds": ["UNREAD"]
        }).execute()

if __name__ == "__main__":
    try:
        creds = autenticar_gmail()
        service = build("gmail", "v1", credentials=creds)
        processar_emails(service)
        print("\nProcessamento concluído! Verifique candidatos.xlsx")
    except HttpError as error:
        print(f"Ocorreu um erro: {error}")
